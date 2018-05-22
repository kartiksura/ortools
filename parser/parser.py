from openpyxl import load_workbook
from enum import Enum
from operator import itemgetter
from datetime import datetime, timedelta
from modulos.classError import GenericError
import json


class PesosStd:
    P_No_Disponible = 0  # No planificable
    P_Disponible = 1    # Coste mínimo
    P_Horario = 9  # Penalización por Horario no preferente
    P_Jornada = 49 # Penalización por Jornada no preferente
    P_Puesto = 99 # Pensalización por puesto no preferente
    P_Festivo = 799 # Penalización por planificar en festivo
    P_Maximo = 999 # Penalización máxima


class Disponibilidad(Enum):
    P_No_Disponible = 0  # No planificable
    P_Disponible = 1
    P_No_Preferente = -1


'''
LIMITACIONES ENCONTRADAS:

    - No se pueden asignar horarios preferidos o permitidos por dias de la semana, por ejemplo: el miercoles sólo
        puede realizar horario de MA, pero los demas dias no puede...
        
'''


class JsonDo:


    '''
    Manipulate Json strings, load, save, etc..
    '''

    # def __init__(self):

    def JsonStringlify(self, inputStr=""):


        res = json.dumps( inputStr , separators=(',', ':'))

        return res

    def JsonSaveToArchive(self, lista, filename):


        with open(filename, 'w') as outfile:
            json.dump(lista, outfile, separators=(',', ':'), default=str)

        outfile.close()

    def JsonLoadFromArchive(self, filename):

        res = []
        line = True

        with open(filename, 'r') as inputfile:
            res = json.load(inputfile)

        inputfile.close()

        return res


class ExcelLoad:


    '''
    Manipulate an excel document, load tables, etc...
    '''

    def __init__(self, documento=None):
        
        self.Empleados = []
        self.hoja = None

        if not documento is None:
            wb = load_workbook(filename = documento)
            self.hoja = wb['Empleados']
        
    
    def OpenXlsx(self, documento):

        wb = load_workbook(filename = documento)
        self.hoja = wb['Empleados']


    def maxRows(self):
        
        # ws.min_col, ws.min_row, ws.max_col, ws.max_row 
        return (self.hoja.max_row)

   
    def LoadEmployees(self):


        for fila in range(2, self.maxRows()+1):
            Empleado = self._LoadEmployee(fila)
            self.Empleados.append(Empleado)

        return self.Empleados


    def _AdaptarLista(self, lista):

        res = []
        _lista = lista.split(',')

        for elem in _lista:
            s = elem.strip()
            res.append(s)

        return res


    def _LoadEmployee(self, fila):

        Empleado = {}
        JornadaSemanal = []

        # Lectura de la hoja excel de empleados y sus características

        Empleado['Id'] = self.hoja.cell(row=fila, column=1).value
        Empleado['Nombre'] = self.hoja.cell(row=fila, column=2).value
        Empleado['Prioridad'] = self.hoja.cell(row=fila, column=3).value
        Empleado['HorasAcumuladas'] = self.hoja.cell(row=fila, column=4).value
        Empleado['HorasContrato'] = self.hoja.cell(row=fila, column=5).value
        Empleado['VacDisfrutadas'] = self.hoja.cell(row=fila, column=6).value
        Empleado['VacContrato'] = self.hoja.cell(row=fila, column=7).value
        Empleado['FestAnterior'] = self.hoja.cell(row=fila, column=8).value
        matriz = self._AdaptarLista(self.hoja.cell(row=fila, column=9).value)
        Empleado['HorPermitidos'] = matriz
        matriz = self._AdaptarLista(self.hoja.cell(row=fila, column=10).value)
        Empleado['HorPreferidos'] = matriz
        matriz = self._AdaptarLista(self.hoja.cell(row=fila, column=11).value)
        Empleado['PuestosPermitidos'] = matriz
        matriz = self._AdaptarLista(self.hoja.cell(row=fila, column=12).value)
        Empleado['PuestosAfinidad'] = matriz
        Empleado['FechaFinContrato'] = self.hoja.cell(row=fila, column=13).value

        # Cargamos la jornada semanal
        for d in range(7):
            ecol = d + 14
            JornadaSemanal.append( self.hoja.cell(row=fila, column=ecol).value)
        Empleado['JornadaSemanal'] = JornadaSemanal

        return Empleado


class Orientador:


    '''
    La misión de esta clase es suministrar el orden y el filtrado de los empleados a entrar en el motor de
    cálculo.
    '''

    def __init__(self, Empleados):


        self.Empleados = Empleados
        self.DispJ = Disponibilidad
        self.Pesos = None
        self.calendario = []
        self.fechaInicial = None
        self.fechaFinal  = None
        self.diferencia = None


    def cargaEmpleados(self, Empleados):


        self.Empleados = Empleados


    def _setJornada(self, diaSemana, tablaDisponibilidad):

        # TablaDisponibilidades:
        # P_No_Disponible = 0  # No planificable
        # P_Disponible = 1
        # P_No_Preferente = -1

        pesoJ = tablaDisponibilidad[diaSemana]
        # print ('Dia semana %i, peso = %i' %(diaSemana, pesoJ))

        if pesoJ == self.DispJ.P_No_Preferente.value:
            pesoJ = self.Pesos.P_Jornada

        return pesoJ


    def calculaPesos(self, FechaInicial, FechaFinal, TablaDePesos):
        '''
        Asigna pesos a los empleados cada dia del calendario, segun la tabla de pesos especificada
        Importante: La matriz de entrada de empleados, ya debe estar filtrada por un solo puesto y
        los horarios permitidos.
        :return:
        '''


        self.fechaInicial = FechaInicial
        self.fechaFinal = FechaFinal
        self.Pesos = TablaDePesos
        self.diferencia = FechaFinal - FechaInicial
        if self.diferencia.days < 0:
            raise GenericError('Las fechas introducidas no son válidas', 'No se puede continuar')

        # calculamos los pesos segun la disponibilidad semanal

        for empleado in self.Empleados:
            plan = {}
            for d in range(self.diferencia.days+1):
                dia = self.fechaInicial + timedelta(days=d)
                diaf = dia.__format__('%Y-%m-%d')
                plan['Id'] = empleado['Id']
                plan['Nombre'] = empleado['Nombre']

                # Pesos definidos segun jornada laboral
                plan[str(diaf)] = self._setJornada( dia.weekday(), empleado['JornadaSemanal'])

                # Si el dia de trabajo esta permitido, entonces, calcula pesos..
                if int(plan[str(diaf)]) != 0:

                    # Asignar pesos de los horarios
                    horario = empleado['HorPermitidos']
                    if not horario in empleado['HorPreferidos']:
                        plan[str(diaf)] =  int(plan[str(diaf)]) + int(self.Pesos.P_Horario)

                    # Asignar pesos segun puestos
                    afinidad = int(empleado['PuestosAfinidad'])
                    if afinidad != 100:
                        calculo = round(afinidad * int(self.Pesos.P_Puesto) / 100,0)
                        plan[str(diaf)] = int(plan[str(diaf)]) + int(calculo)


            self.calendario.append(plan)

        return self.calendario


    def filtrarEmpPorPuesto(self, Puesto, Empleados = None):


        '''
        Filtra el diccionario de empleados segun el puesto en concreto
        :return: Una lista de empleados que coinciden con el puesto solicitado
        '''

        res = []
        ListaEmpleados =  Empleados

        if Empleados is None:
            ListaEmpleados = self.Empleados

        for empleado in ListaEmpleados:
            if Puesto in empleado['PuestosPermitidos']:
                idx = empleado['PuestosPermitidos'].index(Puesto)
                empleado['PuestosAfinidad'] = empleado['PuestosAfinidad'][idx]
                res.append(empleado)

        # Ordenamos por prioridad y luego por afinidad
        ordenado_afi = sorted(res, key=itemgetter('PuestosAfinidad'), reverse=True)
        ordenado_pri = sorted(ordenado_afi, key=itemgetter('Prioridad'))

        return ordenado_pri

    def filtrarEmpPorHorario(self, Horario, Empleados = None):


        '''
        Filtra el diccionario de empleados segun el horario en concreto
        :return: Una lista de empleados que coinciden con el horario solicitado
        '''

        res = []
        pref = []
        ListaEmpleados =  Empleados

        if Empleados is None:
            ListaEmpleados = self.Empleados

        for empleado in ListaEmpleados:
            if Horario in empleado['HorPermitidos']:
                empleado['HorPermitidos'] = Horario
                res.append(empleado)

        temp = res.copy()

        for empleado in res:
            if Horario in empleado['HorPreferidos']:
                pref.append(empleado)
                temp.remove(empleado)

        res = pref + temp

        return res


def test1():


    jsonfile = 'prueba.json'

    lx = ExcelLoad('Empleados.xlsx')
    js = JsonDo()

    res = lx.LoadEmployees()

    for idx, emp in enumerate(res, start=1):
        print('RES,#' + str(idx) + ' ' + str(emp))

    # Pasamos el orientador para devolver una lista de empleados:
    # Que puedan cubrir el puesto en el turno especificado y ordenado,
    # por :
    # 1.- Prioridad
    # 2.- Preferencia de horario

    orientador = Orientador(res)
    resFiltro = orientador.filtrarEmpPorHorario('MA')
    resFiltro = orientador.filtrarEmpPorPuesto('Operario', resFiltro)

    for idx, emp in enumerate(resFiltro, start=1):
        print('#' + str(idx) + ' ' + str(emp))

    js.JsonSaveToArchive(resFiltro, jsonfile)
    #loaded = js.JsonLoadFromArchive(jsonfile)

    #for idx, emp in enumerate(loaded, start=1):
    #    print('R' + str(idx) + ' ' + str(emp))


def test2():

    jsonfile = 'prueba.json'

    js = JsonDo()


    loaded = js.JsonLoadFromArchive(jsonfile)

    for idx, emp in enumerate(loaded, start=1):
        print('Loaded #' + str(idx) + ' ' + str(emp))

    # -------------------

    fini = datetime(2018, 1, 1)
    ffin = datetime(2018, 1, 7)

    orientador = Orientador(loaded)


    res = orientador.calculaPesos(fini, ffin, PesosStd)

    print (str(res))


def main():

    #test1()
    test2()



if __name__ == "__main__":
    main()